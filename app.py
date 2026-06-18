from flask import Flask, jsonify, request, send_from_directory, session, send_file
from functools import wraps
from datetime import datetime, timedelta, timezone
from sqlalchemy.exc import IntegrityError
import os, threading

from models import db, User, Room, RoomMember, RoomHolding, RoomTransaction, Deposit
from stock_service import get_room_service, cleanup_room_service, STOCKS, SECTORS
from education_data import GLOSSARY, GUIDES, TIPS, QUIZ_QUESTIONS
import time, random as _random

app = Flask(__name__, static_folder='static')
app.secret_key = os.environ.get('SECRET_KEY', 'mock-stock-game-secret-2024')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///game.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

def _set_sqlite_pragmas(dbapi_conn, _):
    cur = dbapi_conn.cursor()
    cur.execute("PRAGMA journal_mode=WAL")
    cur.execute("PRAGMA synchronous=NORMAL")
    cur.execute("PRAGMA busy_timeout=5000")
    cur.execute("PRAGMA cache_size=-32000")
    cur.close()

with app.app_context():
    db.create_all()
    from sqlalchemy import event as _sa_event
    _sa_event.listen(db.engine, "connect", _set_sqlite_pragmas)

# ── Server-side room cache (1.5s TTL) ────────────────────────
_room_cache: dict = {}
_room_cache_lock = threading.Lock()
ROOM_CACHE_TTL = 1.5

def _invalidate_room_cache(rid):
    with _room_cache_lock:
        _room_cache.pop(rid, None)

def _get_room_cached(room, uid):
    rid = room.id
    now = time.time()
    with _room_cache_lock:
        entry = _room_cache.get(rid)
    if entry and now - entry['ts'] < ROOM_CACHE_TTL:
        d = dict(entry['data'])
        d['is_host'] = uid == room.host_id
        return d
    data = room_dict(room, uid)
    base = {k: v for k, v in data.items() if k != 'is_host'}
    with _room_cache_lock:
        _room_cache[rid] = {'ts': now, 'data': base}
    return data

# ── News cache (2s TTL) ───────────────────────────────────────
_news_cache: dict = {}
_news_cache_lock = threading.Lock()
NEWS_CACHE_TTL = 2.0

def _get_news_cached(rid):
    now = time.time()
    with _news_cache_lock:
        entry = _news_cache.get(rid)
    if entry and now - entry['ts'] < NEWS_CACHE_TTL:
        return entry['data']
    data = get_room_service(rid).get_news()
    with _news_cache_lock:
        _news_cache[rid] = {'ts': now, 'data': data}
    return data

def _invalidate_news_cache(rid):
    with _news_cache_lock:
        _news_cache.pop(rid, None)

# ── Lottery / roulette thread locks ──────────────────────────
_lottery_lock = threading.Lock()
_rlt_lock = threading.Lock()

KST = timezone(timedelta(hours=9))

# ── Helpers ───────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def deco(*a, **kw):
        if 'user_id' not in session:
            return jsonify({'error': '로그인이 필요합니다.'}), 401
        return f(*a, **kw)
    return deco

def cur_user():
    return db.session.get(User, session['user_id'])

def member_total_value(rid, uid):
    member = RoomMember.query.filter_by(room_id=rid, user_id=uid).first()
    if not member: return 0
    svc = get_room_service(rid)
    total = member.cash
    for h in RoomHolding.query.filter_by(room_id=rid, user_id=uid).all():
        if h.shares > 0:
            p = svc.get_price(h.symbol)
            if p: total += p * h.shares
    for d in Deposit.query.filter_by(room_id=rid, user_id=uid, status='active').all():
        total += d.amount
    return total

def _end_room(room):
    room.status = 'ended'
    now = datetime.utcnow()
    total_seconds = room.duration_minutes * 60
    if room.paused_at:
        game_end = room.paused_at
    elif room.end_time:
        game_end = min(now, room.end_time)
    else:
        game_end = now
    # 예금 이자 정산
    member_map = {m.user_id: m for m in RoomMember.query.filter_by(room_id=room.id).all()}
    for d in Deposit.query.filter_by(room_id=room.id, status='active').all():
        m = member_map.get(d.user_id)
        if m:
            held_seconds = max(0.0, (game_end - d.created_at).total_seconds())
            ratio = min(1.0, held_seconds / total_seconds) if total_seconds > 0 else 1.0
            interest = round(d.amount * d.rate / 100 * ratio, 0)
            d.interest_earned = interest
            d.status = 'matured'
            m.cash += d.amount + interest
    # 보유 주식을 현재가로 현금 청산 (StockService 삭제 전에 실행)
    svc = get_room_service(room.id)
    for h in RoomHolding.query.filter_by(room_id=room.id).all():
        if h.shares > 0:
            price = svc.get_price(h.symbol) or h.avg_price
            m = member_map.get(h.user_id)
            if m:
                m.cash += price * h.shares
        db.session.delete(h)
    db.session.commit()
    cleanup_room_service(room.id)
    _lots.pop(room.id, None)
    _rlt_active.pop(room.id, None)
    _quiz_settings.pop(room.id, None)
    _roulette_config.pop(room.id, None)
    for k in [k for k in _quiz_state if k[0] == room.id]:
        del _quiz_state[k]
    _invalidate_room_cache(room.id)
    _invalidate_news_cache(room.id)

# ── Lottery System ────────────────────────────────────────
_lots = {}   # room_id -> {done: set(), current: dict|None}

LOTTO_PICK_SECS = 60   # participant number-picking window
LOTTO_DRAW_SECS = 45   # host number-drawing window

def _lot_round_due(room, remaining, total_s):
    if room.status != 'active' or total_s <= 0: return None
    lot = _lots.get(room.id, {})
    cur = lot.get('current')
    if cur and cur.get('state') in ('picking', 'drawing'): return None
    done = lot.get('done', set())
    pct = 1 - remaining / total_s  # fraction elapsed
    if pct >= 1/3 and pct < 2/3 and 1 not in done: return 1
    if pct >= 2/3 and 2 not in done: return 2
    return None

def _do_reveal(rid, cur):
    winning = cur['winning']
    results = {}
    for uid_str, picks in cur.get('picks', {}).items():
        uid = int(uid_str)
        matched = len(set(picks) & set(winning))
        prize = (cur['prize'] if matched == 6
                 else round(cur['prize'] / 25) if matched == 5
                 else round(cur['prize'] / 18) if matched == 4
                 else round(cur['prize'] / 11) if matched == 3
                 else 0)
        results[uid_str] = {'matched': matched, 'prize': prize, 'picks': picks}
        if prize > 0:
            m = RoomMember.query.filter_by(room_id=rid, user_id=uid).first()
            if m:
                m.cash += prize
                db.session.add(RoomTransaction(room_id=rid, user_id=uid, symbol='LOTTO',
                    action='ADJ', shares=0, price=0, amount=prize,
                    note=f'복권 {matched}개 일치'))
    db.session.commit()
    cur['results'] = results
    cur['state'] = 'revealed'
    _lots.setdefault(rid, {}).setdefault('done', set()).add(cur['round'])
    # Auto-resume the game if we paused it for the lottery
    if cur.get('auto_paused'):
        room = db.session.get(Room, rid)
        if room and room.status == 'paused' and room.paused_at:
            now_dt = datetime.utcnow()
            room.end_time += (now_dt - room.paused_at)
            room.status = 'active'
            room.paused_at = None
            db.session.commit()
        cur['auto_paused'] = False

ROULETTE_OUTCOMES = [
    {'label': '꽝',  'multiplier': 0,  'weight': 70, 'seg_start': 0,     'seg_end': 252},
    {'label': '1배', 'multiplier': 1,  'weight': 20, 'seg_start': 252,   'seg_end': 324},
    {'label': '2배', 'multiplier': 2,  'weight': 7,  'seg_start': 324,   'seg_end': 349.2},
    {'label': '5배', 'multiplier': 5,  'weight': 2,  'seg_start': 349.2, 'seg_end': 356.4},
    {'label': '25배','multiplier': 25, 'weight': 1,  'seg_start': 356.4, 'seg_end': 360},
]

_roulette_config = {}  # room_id -> {'multipliers': [...], 'weights': [...]}
_DEFAULT_RLT_CFG = {'multipliers': [0, 1, 2, 5, 25], 'weights': [70, 20, 7, 2, 1]}
_rlt_active = {}  # room_id -> {'count': int, 'auto_paused': bool}

def _rlt_cfg(rid):
    c = _roulette_config.get(rid, {})
    return (
        c.get('multipliers', list(_DEFAULT_RLT_CFG['multipliers'])),
        c.get('weights',     list(_DEFAULT_RLT_CFG['weights'])),
    )

def _rlt_outcomes(rid):
    mults, weights = _rlt_cfg(rid)
    total = sum(weights) or 1
    cumulative = 0.0
    result = []
    for i, o in enumerate(ROULETTE_OUTCOMES):
        m = mults[i] if i < len(mults) else o['multiplier']
        w = weights[i] if i < len(weights) else o['weight']
        pct = w / total
        seg_start = round(cumulative * 360, 4)
        seg_end   = round((cumulative + pct) * 360, 4)
        cumulative += pct
        label = '꽝' if m == 0 else (f'{int(m)}배' if float(m) == int(m) else f'{m}배')
        result.append({'label': label, 'multiplier': m, 'weight': w,
                       'seg_start': seg_start, 'seg_end': seg_end})
    return result

def room_dict(room, uid=None):
    now = datetime.utcnow()
    remaining = 0
    if room.status in ('active', 'paused') and room.end_time:
        if room.status == 'paused' and room.paused_at:
            remaining = max(0, int((room.end_time - room.paused_at).total_seconds()))
        else:
            remaining = max(0, int((room.end_time - now).total_seconds()))
    host = db.session.get(User, room.host_id)
    total_s = room.duration_minutes * 60
    return {
        'id': room.id, 'name': room.name, 'code': room.code,
        'host_id': room.host_id, 'host_name': host.username if host else '',
        'status': room.status,
        'duration_minutes': room.duration_minutes,
        'starting_cash': room.starting_cash,
        'deposit_rate': room.deposit_rate,
        'remaining_seconds': remaining,
        'end_time': room.end_time.strftime('%Y-%m-%dT%H:%M:%SZ') if room.end_time else None,
        'member_count': RoomMember.query.filter_by(room_id=room.id).count(),
        'is_host': uid == room.host_id,
        'minigame_available': room.status == 'active' and total_s > 0 and remaining <= int(total_s * 0.05),
        'lottery_round_due': _lot_round_due(room, remaining, total_s) if room.status == 'active' else None,
        'lottery_active': (_lots.get(room.id, {}).get('current') or {}).get('state') in ('picking', 'drawing', 'revealed'),
        'lottery_current_round': (_lots.get(room.id, {}).get('current') or {}).get('round'),
    }

def find_active_room(uid):
    r = Room.query.filter(Room.host_id == uid, Room.status.in_(['waiting','active','paused'])).first()
    if r: return r
    m = RoomMember.query.join(Room).filter(
        RoomMember.user_id == uid, Room.status.in_(['waiting','active','paused'])
    ).first()
    return db.session.get(Room, m.room_id) if m else None


# ── Static ────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/pomodoro')
def pomodoro():
    return send_from_directory('static', 'pomodoro.html')


# ── Auth ──────────────────────────────────────────────────

@app.route('/api/auth/enter', methods=['POST'])
def enter():
    d = request.json or {}
    u = d.get('username', '').strip()
    if not u or len(u) < 2 or len(u) > 30:
        return jsonify({'error': '닉네임은 2~20자 사이여야 합니다.'}), 400
    user = User.query.filter_by(username=u).first()
    if not user:
        user = User(username=u)
        db.session.add(user)
        db.session.commit()
    session['user_id'] = user.id
    ar = find_active_room(user.id)
    return jsonify({'user': user.to_dict(), 'active_room': room_dict(ar, user.id) if ar else None})

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.pop('user_id', None)
    return jsonify({'ok': True})

@app.route('/api/auth/me')
def get_me():
    if 'user_id' not in session:
        return jsonify({'error': 'unauth'}), 401
    user = db.session.get(User, session['user_id'])
    if not user:
        session.pop('user_id', None)
        return jsonify({'error': 'unauth'}), 401
    ar = find_active_room(user.id)
    return jsonify({'user': user.to_dict(), 'active_room': room_dict(ar, user.id) if ar else None})


# ── Room management ───────────────────────────────────────

@app.route('/api/rooms', methods=['POST'])
@login_required
def create_room():
    d = request.json or {}
    name = d.get('name','').strip()
    if not name or len(name) < 2:
        return jsonify({'error': '방 이름은 2자 이상이어야 합니다.'}), 400
    user = cur_user()
    # 2시간 이상 방치된 stale 방 자동 정리
    stale_cutoff = datetime.utcnow() - timedelta(hours=2)
    stale = Room.query.filter(
        Room.host_id == user.id,
        Room.status.in_(['active', 'paused']),
        Room.end_time < stale_cutoff
    ).first()
    if stale:
        _end_room(stale)
    if Room.query.filter(Room.host_id == user.id, Room.status.in_(['waiting','active','paused'])).first():
        return jsonify({'error': '이미 진행 중인 방이 있습니다.'}), 400
    room = Room(
        name=name, host_id=user.id,
        duration_minutes=max(1, min(180, int(d.get('duration_minutes', 30)))),
        starting_cash=max(100000, float(d.get('starting_cash', 10_000_000))),
        deposit_rate=max(0, min(50, float(d.get('deposit_rate', 3.0)))),
    )
    db.session.add(room)
    db.session.commit()
    return jsonify({'room': room_dict(room, user.id)})

@app.route('/api/rooms/join', methods=['POST'])
@login_required
def join_room():
    code = (request.json or {}).get('code','').strip().upper()
    room = Room.query.filter_by(code=code).first()
    if not room: return jsonify({'error': '유효하지 않은 방 코드입니다.'}), 404
    if room.status == 'ended': return jsonify({'error': '이미 종료된 방입니다.'}), 400
    user = cur_user()
    if room.host_id != user.id and not RoomMember.query.filter_by(room_id=room.id, user_id=user.id).first():
        try:
            db.session.add(RoomMember(room_id=room.id, user_id=user.id, cash=room.starting_cash))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
    return jsonify({'room': room_dict(room, user.id)})

def _auto_start_lottery_if_due(room):
    if room.status != 'active' or not room.end_time: return
    now = datetime.utcnow()
    remaining = max(0, int((room.end_time - now).total_seconds()))
    total_s = room.duration_minutes * 60
    round_due = _lot_round_due(room, remaining, total_s)
    if not round_due: return
    with _lottery_lock:
        lot = _lots.setdefault(room.id, {'done': set()})
        if (lot.get('current') or {}).get('state') in ('picking', 'drawing', 'revealed'): return
        member_count = RoomMember.query.filter_by(room_id=room.id).count()
        default_prize = member_count * 100_000_000
        room.status = 'paused'
        room.paused_at = now
        db.session.commit()
        lot['current'] = {
            'round': round_due, 'state': 'picking', 'prize': default_prize,
            'picks': {}, 'winning': None, 'results': None,
            'pick_dl': time.time() + LOTTO_PICK_SECS, 'draw_dl': None,
            'auto_paused': True,
        }
    _invalidate_room_cache(room.id)

@app.route('/api/rooms/<int:rid>')
@login_required
def get_room(rid):
    room = Room.query.get_or_404(rid)
    now_dt = datetime.utcnow()
    if room.status == 'active' and room.end_time and now_dt >= room.end_time:
        _end_room(room)
        return jsonify(room_dict(room, cur_user().id))
    # paused 상태로 고착된 채 end_time 초과 시에도 자동 종료
    if room.status == 'paused' and room.paused_at and room.end_time:
        if (room.end_time - room.paused_at).total_seconds() <= 0:
            _end_room(room)
            return jsonify(room_dict(room, cur_user().id))
    prev_status = room.status
    _auto_start_lottery_if_due(room)
    if room.status != prev_status:
        _invalidate_room_cache(rid)
    return jsonify(_get_room_cached(room, cur_user().id))

@app.route('/api/rooms/<int:rid>/start', methods=['POST'])
@login_required
def start_room(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '진행자만 게임을 시작할 수 있습니다.'}), 403
    if room.status != 'waiting': return jsonify({'error': '이미 시작되었거나 종료된 방입니다.'}), 400
    now = datetime.utcnow()
    room.status = 'active'
    room.start_time = now
    room.end_time = now + timedelta(minutes=room.duration_minutes)
    db.session.commit()
    _invalidate_room_cache(rid)
    return jsonify({'room': room_dict(room, user.id)})

@app.route('/api/rooms/<int:rid>/pause', methods=['POST'])
@login_required
def pause_room(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '진행자만 일시정지할 수 있습니다.'}), 403
    if room.status != 'active': return jsonify({'error': '진행 중인 게임만 일시정지할 수 있습니다.'}), 400
    room.status = 'paused'
    room.paused_at = datetime.utcnow()
    db.session.commit()
    _invalidate_room_cache(rid)
    return jsonify({'room': room_dict(room, user.id)})

@app.route('/api/rooms/<int:rid>/resume', methods=['POST'])
@login_required
def resume_room(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '진행자만 재개할 수 있습니다.'}), 403
    if room.status != 'paused': return jsonify({'error': '일시정지된 게임만 재개할 수 있습니다.'}), 400
    now = datetime.utcnow()
    paused_duration = now - (room.paused_at or now)
    room.end_time += paused_duration
    room.status = 'active'
    room.paused_at = None
    db.session.commit()
    _invalidate_room_cache(rid)
    return jsonify({'room': room_dict(room, user.id)})

@app.route('/api/rooms/<int:rid>/end', methods=['POST'])
@login_required
def end_room(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '진행자만 종료할 수 있습니다.'}), 403
    if room.status == 'ended': return jsonify({'error': '이미 종료된 방입니다.'}), 400
    _end_room(room)
    return jsonify({'room': room_dict(room, user.id)})


# ── Host endpoints ────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/host/members')
@login_required
def host_members(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    members = RoomMember.query.filter_by(room_id=rid).all()
    uids = [m.user_id for m in members]
    user_map = {u.id: u for u in db.session.query(User).filter(User.id.in_(uids)).all()}
    result = []
    for m in members:
        u = user_map.get(m.user_id)
        total = member_total_value(rid, m.user_id)
        gain_pct = (total - room.starting_cash) / room.starting_cash * 100 if room.starting_cash else 0
        result.append({
            'user_id': m.user_id, 'username': u.username if u else str(m.user_id),
            'cash': m.cash, 'total_value': round(total,0), 'gain_pct': round(gain_pct,2),
        })
    result.sort(key=lambda x: x['total_value'], reverse=True)
    for i, r in enumerate(result): r['rank'] = i + 1
    return jsonify(result)

@app.route('/api/rooms/<int:rid>/host/members/<int:uid>', methods=['DELETE'])
@login_required
def kick_member(rid, uid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    if room.status != 'waiting': return jsonify({'error': '대기 중인 방에서만 강퇴할 수 있습니다.'}), 400
    m = RoomMember.query.filter_by(room_id=rid, user_id=uid).first()
    if not m: return jsonify({'error': '참여자를 찾을 수 없습니다.'}), 404
    db.session.delete(m)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/rooms/<int:rid>/host/lobby-members')
@login_required
def lobby_members(rid):
    Room.query.get_or_404(rid)
    result = []
    for m in RoomMember.query.filter_by(room_id=rid).all():
        u = db.session.get(User, m.user_id)
        result.append({'user_id': m.user_id, 'username': u.username if u else str(m.user_id)})
    return jsonify(result)

@app.route('/api/rooms/<int:rid>/host/adjust', methods=['POST'])
@login_required
def host_adjust(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    d = request.json or {}
    target_uid = d.get('user_id')
    delta = float(d.get('delta', 0))
    note = d.get('note', '진행자 자산 조정')
    m = RoomMember.query.filter_by(room_id=rid, user_id=target_uid).first()
    if not m: return jsonify({'error': '참여자를 찾을 수 없습니다.'}), 404
    m.cash = max(0, m.cash + delta)
    db.session.add(RoomTransaction(room_id=rid, user_id=target_uid, symbol='ADJ', action='ADJ', amount=delta, note=note))
    db.session.commit()
    target = db.session.get(User, target_uid)
    return jsonify({'message': f'{target.username} 자산 {delta:+,.0f}원 조정', 'new_cash': m.cash})


@app.route('/api/rooms/<int:rid>/host/members/<int:uid>/transactions')
@login_required
def host_member_transactions(rid, uid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    page = request.args.get('page', 1, type=int)
    pg = (RoomTransaction.query.filter_by(room_id=rid, user_id=uid)
          .order_by(RoomTransaction.timestamp.desc())
          .paginate(page=page, per_page=30, error_out=False))
    target = db.session.get(User, uid)
    return jsonify({
        'username': target.username if target else str(uid),
        'transactions': [{
            'id': t.id,
            'name': STOCKS.get(t.symbol,{}).get('name', '자산조정') if t.action != 'ADJ' else '자산조정',
            'action': t.action, 'shares': t.shares, 'price': t.price,
            'amount': t.amount, 'note': t.note,
            'timestamp': t.timestamp.replace(tzinfo=timezone.utc).astimezone(KST).strftime('%m-%d %H:%M'),
        } for t in pg.items],
        'total': pg.total, 'pages': pg.pages, 'current_page': page,
    })


@app.route('/api/rooms/<int:rid>/host/news-interval', methods=['GET', 'POST'])
@login_required
def host_news_interval(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    svc = get_room_service(rid)
    if request.method == 'POST':
        d = request.json or {}
        if 'news_seconds' in d:
            svc.set_news_interval(float(d['news_seconds']))
        if 'price_seconds' in d:
            svc.set_price_interval(float(d['price_seconds']))
    return jsonify({
        'news_seconds': svc.get_news_interval(),
        'price_seconds': svc.get_price_interval(),
    })


# ── Stocks ────────────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/stocks')
@login_required
def get_stocks(rid):
    Room.query.get_or_404(rid)
    svc = get_room_service(rid)
    sf = request.args.get('sector','')
    result = []
    for sym, info in STOCKS.items():
        if sf and info['sector'] != sf: continue
        price = svc.get_price(sym)
        prev  = svc.get_prev_close(sym)
        if price:
            ch = (price - prev) if prev else 0
            ch_pct = (ch / prev * 100) if prev else 0
            result.append({
                'symbol': sym, 'name': info['name'],
                'sector': info['sector'],
                'price': round(price, 0), 'change': round(ch, 0),
                'change_pct': round(ch_pct, 2),
            })
    return jsonify({'stocks': result, 'sectors': SECTORS})

@app.route('/api/rooms/<int:rid>/host/force-price', methods=['POST'])
@login_required
def host_force_price(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    d = request.json or {}
    symbol = d.get('symbol', '')
    pct = float(d.get('pct', 0))
    if not symbol or abs(pct) > 50:
        return jsonify({'error': '잘못된 요청'}), 400
    new_price = get_room_service(rid).force_price(symbol, pct)
    if new_price is None:
        return jsonify({'error': '종목 없음'}), 404
    return jsonify({'symbol': symbol, 'price': new_price})


@app.route('/api/rooms/<int:rid>/host/send-news', methods=['POST'])
@login_required
def host_send_news(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    d = request.json or {}
    show_hint = bool(d.get('show_hint', True))
    svc = get_room_service(rid)
    svc.trigger_news(show_hint)
    _invalidate_news_cache(rid)
    return jsonify(svc.get_news())

@app.route('/api/rooms/<int:rid>/news')
@login_required
def get_room_news(rid):
    Room.query.get_or_404(rid)
    return jsonify(_get_news_cached(rid))


@app.route('/api/rooms/<int:rid>/stocks/<symbol>/chart')
@login_required
def get_chart(rid, symbol):
    Room.query.get_or_404(rid)
    if symbol not in STOCKS: return jsonify({'error': '종목 없음'}), 404
    pm = {'1d':('1d','5m'),'1w':('5d','30m'),'1mo':('1mo','1d'),'3mo':('3mo','1d'),'1y':('1y','1wk')}
    period = request.args.get('period','1mo')
    yp, yi = pm.get(period, ('1mo','1d'))
    hist = get_room_service(rid).get_history(symbol, period=yp, interval=yi)
    return jsonify({'symbol': symbol, 'name': STOCKS[symbol]['name'], 'history': hist})


# ── Trade ─────────────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/trade', methods=['POST'])
@login_required
def trade(rid):
    room = Room.query.get_or_404(rid)
    if room.status != 'active':
        return jsonify({'error': '게임이 진행 중이 아닙니다.'}), 400
    if room.end_time and datetime.utcnow() >= room.end_time:
        return jsonify({'error': '게임이 종료되었습니다.'}), 400
    user = cur_user()
    member = RoomMember.query.filter_by(room_id=rid, user_id=user.id).first()
    if not member: return jsonify({'error': '이 방의 참여자가 아닙니다.'}), 403
    d = request.json or {}
    symbol = d.get('symbol','')
    action = d.get('action','').upper()
    try: shares = int(d.get('shares', 0))
    except: return jsonify({'error': '수량 오류'}), 400
    if symbol not in STOCKS: return jsonify({'error': '유효하지 않은 종목'}), 400
    if action not in ('BUY','SELL'): return jsonify({'error': '유효하지 않은 거래'}), 400
    if shares <= 0: return jsonify({'error': '수량은 1 이상'}), 400
    price = get_room_service(rid).get_price(symbol)
    if not price: return jsonify({'error': '주가를 불러올 수 없습니다.'}), 500
    holding = RoomHolding.query.filter_by(room_id=rid, user_id=user.id, symbol=symbol).first()
    amount = price * shares
    if action == 'BUY':
        if member.cash < amount:
            return jsonify({'error': f'잔액 부족 — 필요: {amount:,.0f}원 / 보유: {member.cash:,.0f}원'}), 400
        member.cash -= amount
        if holding:
            ns = holding.shares + shares
            holding.avg_price = (holding.avg_price * holding.shares + amount) / ns
            holding.shares = ns
        else:
            db.session.add(RoomHolding(room_id=rid, user_id=user.id, symbol=symbol, shares=shares, avg_price=price))
    else:
        if not holding or holding.shares < shares:
            return jsonify({'error': f'보유 수량 부족 — 보유: {holding.shares if holding else 0}주'}), 400
        member.cash += amount
        holding.shares -= shares
        if holding.shares == 0: db.session.delete(holding)
    db.session.add(RoomTransaction(room_id=rid, user_id=user.id, symbol=symbol,
                                   action=action, shares=shares, price=price, amount=amount))
    db.session.commit()
    return jsonify({'message': f'{STOCKS[symbol]["name"]} {shares}주 {"매수" if action=="BUY" else "매도"} 완료!',
                    'cash': member.cash})


# ── Portfolio ─────────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/portfolio')
@login_required
def get_portfolio(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    m = RoomMember.query.filter_by(room_id=rid, user_id=user.id).first()
    if not m: return jsonify({'error': '참여자가 아닙니다.'}), 403
    svc = get_room_service(rid)
    holdings_data, sv = [], 0.0
    for h in RoomHolding.query.filter_by(room_id=rid, user_id=user.id).all():
        if h.shares <= 0: continue
        price = svc.get_price(h.symbol)
        if not price: continue
        cv = price * h.shares
        gain = cv - h.avg_price * h.shares
        sv += cv
        holdings_data.append({
            'symbol': h.symbol, 'name': STOCKS.get(h.symbol,{}).get('name',h.symbol),
            'sector': STOCKS.get(h.symbol,{}).get('sector',''),
            'shares': h.shares, 'avg_price': h.avg_price, 'current_price': price,
            'current_value': cv, 'gain': gain,
            'gain_pct': round((price - h.avg_price)/h.avg_price*100 if h.avg_price else 0, 2),
        })
    deps_locked = sum(d.amount for d in Deposit.query.filter_by(room_id=rid, user_id=user.id, status='active').all())
    total = m.cash + sv + deps_locked
    start = room.starting_cash
    return jsonify({
        'cash': m.cash, 'stock_value': sv, 'deposits_locked': deps_locked,
        'total_value': total, 'total_gain': total - start,
        'total_gain_pct': round((total - start)/start*100 if start else 0, 2),
        'holdings': sorted(holdings_data, key=lambda x: x['current_value'], reverse=True),
    })


# ── Rankings ──────────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/rankings')
@login_required
def get_rankings(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    start = room.starting_cash
    board = []
    for m in RoomMember.query.filter_by(room_id=rid).all():
        u = db.session.get(User, m.user_id)
        total = member_total_value(rid, m.user_id)
        board.append({'user_id': m.user_id, 'username': u.username,
                      'total_value': round(total,0),
                      'gain_pct': round((total-start)/start*100 if start else 0, 2),
                      'is_me': m.user_id == user.id})
    board.sort(key=lambda x: x['total_value'], reverse=True)
    for i, e in enumerate(board): e['rank'] = i + 1
    return jsonify(board)


# ── Transactions ──────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/transactions')
@login_required
def get_transactions(rid):
    Room.query.get_or_404(rid)
    user = cur_user()
    page = request.args.get('page', 1, type=int)
    pg = (RoomTransaction.query.filter_by(room_id=rid, user_id=user.id)
          .order_by(RoomTransaction.timestamp.desc())
          .paginate(page=page, per_page=20, error_out=False))
    return jsonify({
        'transactions': [{
            'id': t.id,
            'name': STOCKS.get(t.symbol,{}).get('name', '자산조정') if t.action != 'ADJ' else '자산조정',
            'action': t.action, 'shares': t.shares, 'price': t.price,
            'amount': t.amount, 'note': t.note,
            'timestamp': t.timestamp.replace(tzinfo=timezone.utc).astimezone(KST).strftime('%m-%d %H:%M'),
        } for t in pg.items],
        'total': pg.total, 'pages': pg.pages, 'current_page': page,
    })


# ── Deposits ──────────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/deposits', methods=['GET'])
@login_required
def get_deposits(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    deps = Deposit.query.filter_by(room_id=rid, user_id=user.id).order_by(Deposit.created_at.desc()).all()
    now = datetime.utcnow()
    total_seconds = room.duration_minutes * 60
    result = []
    for d in deps:
        max_interest = d.amount * d.rate / 100
        if d.status == 'active' and total_seconds > 0:
            held = max(0.0, (now - d.created_at).total_seconds())
            ratio = min(1.0, held / total_seconds)
            expected_interest = round(max_interest * ratio, 0)
        else:
            expected_interest = d.interest_earned or 0
        result.append({
            'id': d.id, 'amount': d.amount, 'rate': d.rate, 'status': d.status,
            'interest_earned': d.interest_earned,
            'expected_interest': expected_interest,
            'max_interest': round(max_interest, 0),
            'created_at': d.created_at.replace(tzinfo=timezone.utc).astimezone(KST).strftime('%m-%d %H:%M'),
        })
    return jsonify(result)

@app.route('/api/rooms/<int:rid>/deposits', methods=['POST'])
@login_required
def create_deposit(rid):
    room = Room.query.get_or_404(rid)
    if room.status != 'active': return jsonify({'error': '게임이 진행 중이 아닙니다.'}), 400
    if room.end_time and datetime.utcnow() >= room.end_time: return jsonify({'error': '게임 종료'}), 400
    user = cur_user()
    m = RoomMember.query.filter_by(room_id=rid, user_id=user.id).first()
    if not m: return jsonify({'error': '참여자가 아닙니다.'}), 403
    try: amount = float((request.json or {}).get('amount', 0))
    except: return jsonify({'error': '금액 오류'}), 400
    if not (0 < amount < float('inf')): return jsonify({'error': '금액은 0보다 커야 합니다.'}), 400
    if m.cash < amount: return jsonify({'error': f'잔액 부족 — 보유: {m.cash:,.0f}원'}), 400
    m.cash -= amount
    dep = Deposit(room_id=rid, user_id=user.id, amount=amount, rate=room.deposit_rate)
    db.session.add(dep)
    db.session.commit()
    total_seconds = room.duration_minutes * 60
    remaining = max(0.0, (room.end_time - datetime.utcnow()).total_seconds()) if room.end_time else total_seconds
    ratio = min(1.0, remaining / total_seconds) if total_seconds > 0 else 1.0
    expected_interest = round(amount * room.deposit_rate / 100 * ratio, 0)
    max_interest = round(amount * room.deposit_rate / 100, 0)
    return jsonify({'message': f'{amount:,.0f}원 예금! 예상 이자 {expected_interest:,.0f}원 (보유 시간 비례 지급).',
                    'cash': m.cash, 'deposit': {'id': dep.id, 'amount': dep.amount,
                    'expected_interest': expected_interest, 'max_interest': max_interest, 'rate': dep.rate}})

@app.route('/api/rooms/<int:rid>/deposits/<int:did>', methods=['DELETE'])
@login_required
def withdraw_deposit(rid, did):
    dep = db.session.get(Deposit, did)
    if not dep: return jsonify({'error': '예금을 찾을 수 없습니다.'}), 404
    user = cur_user()
    if dep.room_id != rid or dep.user_id != user.id: return jsonify({'error': '권한 없음'}), 403
    if dep.status != 'active': return jsonify({'error': '이미 처리된 예금'}), 400
    m = RoomMember.query.filter_by(room_id=rid, user_id=user.id).first()
    dep.status = 'withdrawn'
    m.cash += dep.amount
    db.session.commit()
    return jsonify({'message': f'예금 해지 — {dep.amount:,.0f}원 반환 (이자 없음)', 'cash': m.cash})


# ── Mini-game: Roulette ───────────────────────────────────

@app.route('/api/rooms/<int:rid>/minigame')
@login_required
def get_minigame(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id == user.id:
        return jsonify({'error': '진행자는 참여할 수 없습니다.'}), 403
    m = RoomMember.query.filter_by(room_id=rid, user_id=user.id).first()
    if not m:
        return jsonify({'error': '참여자가 아닙니다.'}), 403
    spins_used = RoomTransaction.query.filter_by(room_id=rid, user_id=user.id, action='RLT').count()
    mults, weights = _rlt_cfg(rid)
    return jsonify({'spins_left': max(0, 3 - spins_used), 'cash': m.cash,
                    'multipliers': mults, 'weights': weights})

@app.route('/api/rooms/<int:rid>/minigame/open', methods=['POST'])
@login_required
def minigame_open(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id == user.id:
        return jsonify({'error': '진행자는 참여할 수 없습니다.'}), 403
    if not RoomMember.query.filter_by(room_id=rid, user_id=user.id).first():
        return jsonify({'error': '참여자가 아닙니다.'}), 403
    paused_now = False
    with _rlt_lock:
        state = _rlt_active.setdefault(rid, {'count': 0, 'auto_paused': False})
        state['count'] += 1
        if state['count'] == 1 and room.status == 'active':
            room.status = 'paused'
            room.paused_at = datetime.utcnow()
            db.session.commit()
            state['auto_paused'] = True
            paused_now = True
    if paused_now:
        _invalidate_room_cache(rid)
    remaining = 0
    if room.paused_at and room.end_time:
        remaining = max(0, int((room.end_time - room.paused_at).total_seconds()))
    return jsonify({'ok': True, 'paused': paused_now, 'remaining_seconds': remaining})

@app.route('/api/rooms/<int:rid>/minigame/close', methods=['POST'])
@login_required
def minigame_close(rid):
    resumed = False
    new_end_time = None
    with _rlt_lock:
        state = _rlt_active.get(rid)
        if not state or state['count'] <= 0:
            return jsonify({'ok': True})
        state['count'] = max(0, state['count'] - 1)
        if state['count'] == 0 and state.get('auto_paused'):
            room = Room.query.get(rid)
            if room and room.status == 'paused' and room.paused_at:
                room.end_time += (datetime.utcnow() - room.paused_at)
                room.status = 'active'
                room.paused_at = None
                db.session.commit()
                resumed = True
                new_end_time = room.end_time.strftime('%Y-%m-%dT%H:%M:%SZ')
            state['auto_paused'] = False
    if resumed:
        _invalidate_room_cache(rid)
    return jsonify({'ok': True, 'resumed': resumed, 'end_time': new_end_time})

@app.route('/api/rooms/<int:rid>/minigame/spin', methods=['POST'])
@login_required
def minigame_spin(rid):
    room = Room.query.get_or_404(rid)
    rlt_paused = _rlt_active.get(rid, {}).get('auto_paused', False)
    if room.status not in ('active', 'paused') or (room.status == 'paused' and not rlt_paused):
        return jsonify({'error': '게임이 진행 중이 아닙니다.'}), 400
    now = datetime.utcnow()
    if room.status == 'paused' and room.paused_at:
        remaining = max(0, (room.end_time - room.paused_at).total_seconds())
    else:
        remaining = max(0, (room.end_time - now).total_seconds()) if room.end_time else 0
    total_s = room.duration_minutes * 60
    if total_s > 0 and remaining > total_s * 0.05:
        return jsonify({'error': '아직 미니게임 시간이 아닙니다.'}), 400
    user = cur_user()
    if room.host_id == user.id:
        return jsonify({'error': '진행자는 참여할 수 없습니다.'}), 403
    m = RoomMember.query.filter_by(room_id=rid, user_id=user.id).first()
    if not m:
        return jsonify({'error': '참여자가 아닙니다.'}), 403
    spins_used = RoomTransaction.query.filter_by(room_id=rid, user_id=user.id, action='RLT').count()
    if spins_used >= 3:
        return jsonify({'error': '기회를 모두 사용했습니다.'}), 400
    try:
        bet = float((request.json or {}).get('bet', 0))
    except (TypeError, ValueError):
        return jsonify({'error': '금액 오류'}), 400
    if bet <= 0 or bet > m.cash:
        return jsonify({'error': '베팅 금액이 올바르지 않습니다.'}), 400
    outcomes = _rlt_outcomes(rid)
    outcome = _random.choices(outcomes, weights=[o['weight'] for o in outcomes])[0]
    winnings = round(bet * outcome['multiplier'])
    net = winnings - bet
    m.cash = m.cash - bet + winnings
    note = f"룰렛 {outcome['label']} (베팅 {bet:,.0f}원)"
    db.session.add(RoomTransaction(room_id=rid, user_id=user.id, symbol='ROULETTE',
                                   action='RLT', shares=0, price=0, amount=net, note=note))
    db.session.commit()
    return jsonify({
        'outcome': outcome['label'], 'multiplier': outcome['multiplier'],
        'bet': bet, 'winnings': winnings, 'net': net, 'cash': m.cash,
        'spins_left': max(0, 3 - (spins_used + 1)),
        'seg_start': outcome['seg_start'], 'seg_end': outcome['seg_end'],
    })


# ── Lottery ───────────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/lottery/start', methods=['POST'])
@login_required
def lottery_start(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '진행자만 시작할 수 있습니다.'}), 403
    if room.status != 'active': return jsonify({'error': '게임이 진행 중이 아닙니다.'}), 400
    d = request.json or {}
    round_n = int(d.get('round', 1))
    prize = float(d.get('prize', 0))
    if prize <= 0: return jsonify({'error': '상금을 입력하세요.'}), 400
    lot = _lots.setdefault(rid, {'done': set()})
    if (lot.get('current') or {}).get('state') in ('picking', 'drawing'):
        return jsonify({'error': '이미 진행 중인 복권이 있습니다.'}), 400
    # Auto-pause the game while lottery runs
    room.status = 'paused'
    room.paused_at = datetime.utcnow()
    db.session.commit()
    lot['current'] = {
        'round': round_n, 'state': 'picking', 'prize': prize,
        'picks': {}, 'winning': None, 'results': None,
        'pick_dl': time.time() + LOTTO_PICK_SECS, 'draw_dl': None,
        'auto_paused': True,
    }
    _invalidate_room_cache(rid)
    return jsonify({'ok': True})

@app.route('/api/rooms/<int:rid>/lottery')
@login_required
def get_lottery(rid):
    lot = _lots.get(rid, {})
    cur = lot.get('current')
    if not cur:
        return jsonify({'state': None})
    now = time.time()
    # 상태 전이는 Lock 내에서 원자적으로 처리 (중복 reveal 방지)
    with _lottery_lock:
        if cur['state'] == 'picking' and now >= cur['pick_dl']:
            cur['state'] = 'drawing'
            cur['draw_dl'] = now + LOTTO_DRAW_SECS
        if cur['state'] == 'drawing' and now >= (cur.get('draw_dl') or float('inf')):
            if not cur.get('winning'):
                cur['winning'] = sorted(_random.sample(range(1, 46), 6))
            _do_reveal(rid, cur)
    user = cur_user()
    uid_str = str(user.id)
    result = {
        'state': cur['state'],
        'round': cur['round'],
        'prize': cur['prize'],
        'pick_deadline': cur.get('pick_dl'),
        'draw_deadline': cur.get('draw_dl'),
        'my_picks': cur.get('picks', {}).get(uid_str),
    }
    if cur['state'] == 'revealed':
        result['winning'] = cur['winning']
        result['my_result'] = (cur.get('results') or {}).get(uid_str)
        room = Room.query.get_or_404(rid)
        if room.host_id == user.id:
            result['all_results'] = cur.get('results', {})
    return jsonify(result)

@app.route('/api/rooms/<int:rid>/lottery/pick', methods=['POST'])
@login_required
def lottery_pick(rid):
    lot = _lots.get(rid, {})
    cur = lot.get('current')
    if not cur or cur.get('state') != 'picking':
        return jsonify({'error': '번호 입력 시간이 아닙니다.'}), 400
    if time.time() >= cur['pick_dl']:
        return jsonify({'error': '입력 시간이 종료되었습니다.'}), 400
    user = cur_user()
    room = Room.query.get_or_404(rid)
    if room.host_id == user.id: return jsonify({'error': '진행자는 참여 불가'}), 403
    if not RoomMember.query.filter_by(room_id=rid, user_id=user.id).first():
        return jsonify({'error': '참여자가 아닙니다.'}), 403
    raw = (request.json or {}).get('numbers', [])
    try:
        nums = sorted(set(int(n) for n in raw if 1 <= int(n) <= 45))
    except Exception:
        return jsonify({'error': '번호 형식 오류'}), 400
    if len(nums) != 6:
        return jsonify({'error': '1~45 사이 숫자 6개를 선택하세요.'}), 400
    cur['picks'][str(user.id)] = nums
    return jsonify({'ok': True, 'picks': nums})

@app.route('/api/rooms/<int:rid>/lottery/draw', methods=['POST'])
@login_required
def lottery_draw(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '진행자만 추첨할 수 있습니다.'}), 403
    lot = _lots.get(rid, {})
    cur = lot.get('current')
    if not cur or cur.get('state') != 'drawing':
        return jsonify({'error': '추첨 시간이 아닙니다.'}), 400
    raw = (request.json or {}).get('numbers', [])
    if raw:
        try:
            nums = sorted(set(int(n) for n in raw if 1 <= int(n) <= 45))
        except Exception:
            return jsonify({'error': '번호 형식 오류'}), 400
        if len(nums) != 6:
            return jsonify({'error': '1~45 사이 숫자 6개를 선택하세요.'}), 400
        cur['winning'] = nums
    else:
        cur['winning'] = sorted(_random.sample(range(1, 46), 6))
    _do_reveal(rid, cur)
    return jsonify({'ok': True, 'winning': cur['winning'], 'results': cur.get('results', {})})

@app.route('/api/rooms/<int:rid>/lottery/skip', methods=['POST'])
@login_required
def lottery_skip(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    d = request.json or {}
    round_n = int(d.get('round', 1))
    lot = _lots.setdefault(rid, {'done': set()})
    lot.setdefault('done', set()).add(round_n)
    return jsonify({'ok': True})


# ── Education ─────────────────────────────────────────────

@app.route('/api/education/glossary')
def get_glossary():
    q = request.args.get('q','').lower()
    return jsonify([g for g in GLOSSARY if not q or q in g['term'].lower() or q in g['definition'].lower()])

@app.route('/api/education/guides')
def get_guides():
    return jsonify(GUIDES)

@app.route('/api/education/guides/<int:gid>')
def get_guide(gid):
    g = next((x for x in GUIDES if x['id'] == gid), None)
    return jsonify(g) if g else (jsonify({'error': '없음'}), 404)

@app.route('/api/education/tips')
def get_tips():
    return jsonify(TIPS)


# ── Quiz ──────────────────────────────────────────────────

_quiz_state: dict = {}
_quiz_settings: dict = {}

@app.route('/api/rooms/<int:rid>/quiz', methods=['GET'])
@login_required
def get_quiz(rid):
    room = Room.query.get_or_404(rid)
    if room.status != 'active':
        return jsonify({'error': '게임 중에만 사용 가능합니다'}), 400
    user = cur_user()
    key = (rid, user.id)
    state = _quiz_state.get(key, {})
    remaining = max(0, state.get('cooldown_until', 0) - time.time())
    if remaining > 0:
        return jsonify({'cooldown': int(remaining)})
    q = _random.choice(QUIZ_QUESTIONS)
    _quiz_state[key] = {'qid': q['id'], 'cooldown_until': 0}
    return jsonify({'id': q['id'], 'question': q['q'], 'cooldown': 0})

@app.route('/api/rooms/<int:rid>/quiz', methods=['POST'])
@login_required
def submit_quiz(rid):
    room = Room.query.get_or_404(rid)
    if room.status != 'active':
        return jsonify({'error': '게임이 종료되었습니다'}), 400
    user = cur_user()
    key = (rid, user.id)
    state = _quiz_state.get(key)
    if not state or state.get('cooldown_until', 0) > time.time():
        return jsonify({'error': '퀴즈를 먼저 시작하세요'}), 400
    d = request.json or {}
    user_answer = bool(d.get('answer'))
    q = next((x for x in QUIZ_QUESTIONS if x['id'] == state['qid']), None)
    if not q:
        return jsonify({'error': '잘못된 요청'}), 400
    correct = user_answer == q['a']
    settings = _quiz_settings.get(rid, {})
    reward_pct = settings.get('reward_pct', 1.0)
    penalty_pct = settings.get('penalty_pct', 0.5)
    reward = 0
    penalty = 0
    member = RoomMember.query.filter_by(room_id=rid, user_id=user.id).first()
    if member:
        if correct:
            reward = max(10000, int(room.starting_cash * reward_pct / 100))
            member.cash += reward
        else:
            penalty = max(5000, int(room.starting_cash * penalty_pct / 100))
            member.cash = max(0, member.cash - penalty)
        db.session.commit()
    _quiz_state[key] = {'qid': None, 'cooldown_until': time.time() + 60}
    return jsonify({'correct': correct, 'reward': reward, 'penalty': penalty, 'explanation': q['ex']})


@app.route('/api/rooms/<int:rid>/host/market-event', methods=['POST'])
@login_required
def host_market_event(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    d = request.json or {}
    sector = d.get('sector', 'all')
    try: pct = float(d.get('pct', 0))
    except: return jsonify({'error': '잘못된 변동률'}), 400
    if pct == 0 or abs(pct) > 50:
        return jsonify({'error': '변동률은 ±1 ~ ±50% 사이'}), 400
    affected = get_room_service(rid).force_sector_event(sector, pct)
    label = '전체 시장' if sector == 'all' else f'{sector} 섹터'
    sign = '+' if pct > 0 else ''
    return jsonify({'message': f'{label} {sign}{pct:.0f}% 적용 ({len(affected)}개 종목)', 'count': len(affected)})


@app.route('/api/rooms/<int:rid>/host/roulette-config', methods=['GET', 'POST'])
@login_required
def host_roulette_config(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id: return jsonify({'error': '권한 없음'}), 403
    if request.method == 'GET':
        mults, weights = _rlt_cfg(rid)
        return jsonify({'multipliers': mults, 'weights': weights})
    d = request.json or {}
    raw_m = d.get('multipliers', _DEFAULT_RLT_CFG['multipliers'])
    raw_w = d.get('weights',     _DEFAULT_RLT_CFG['weights'])
    if not (isinstance(raw_m, list) and isinstance(raw_w, list)
            and len(raw_m) == 5 and len(raw_w) == 5):
        return jsonify({'error': '형식 오류'}), 400
    mults   = [0] + [max(0, float(x)) for x in raw_m[1:]]
    weights = [max(0, float(x)) for x in raw_w]
    if sum(weights) == 0:
        return jsonify({'error': '확률 합계가 0이 될 수 없습니다.'}), 400
    _roulette_config[rid] = {'multipliers': mults, 'weights': weights}
    return jsonify({'ok': True, 'multipliers': mults, 'weights': weights})


@app.route('/api/rooms/<int:rid>/host/quiz-settings', methods=['GET', 'POST'])
@login_required
def quiz_settings(rid):
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id:
        return jsonify({'error': '진행자만 변경할 수 있습니다'}), 403
    if request.method == 'GET':
        s = _quiz_settings.get(rid, {})
        return jsonify({'reward_pct': s.get('reward_pct', 1.0), 'penalty_pct': s.get('penalty_pct', 0.5)})
    d = request.json or {}
    _quiz_settings[rid] = {
        'reward_pct': max(0, min(10, float(d.get('reward_pct', 1.0)))),
        'penalty_pct': max(0, min(10, float(d.get('penalty_pct', 0.5)))),
    }
    return jsonify({'ok': True})


# ── Export ────────────────────────────────────────────────

@app.route('/api/rooms/<int:rid>/export')
@login_required
def export_rankings(rid):
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    room = Room.query.get_or_404(rid)
    user = cur_user()
    if room.host_id != user.id:
        return jsonify({'error': '진행자만 다운로드할 수 있습니다.'}), 403
    start = room.starting_cash
    board = []
    for m in RoomMember.query.filter_by(room_id=rid).all():
        u = db.session.get(User, m.user_id)
        total = member_total_value(rid, m.user_id)
        gain_pct = (total - start) / start * 100 if start else 0
        parts = u.username.split(' ', 1)
        sid  = parts[0] if len(parts) > 1 else ''
        name = parts[1] if len(parts) > 1 else u.username
        board.append({'name': name, 'sid': sid, 'total_value': round(total),
                      'gain_pct': round(gain_pct, 2), 'gain_amount': round(total - start)})
    board.sort(key=lambda x: x['total_value'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '최종 순위'

    thin = Side(style='thin', color='444444')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    up_fill  = PatternFill(start_color='1A3A2A', end_color='1A3A2A', fill_type='solid')
    dn_fill  = PatternFill(start_color='3A1A1A', end_color='3A1A1A', fill_type='solid')
    medal_fills = {1: PatternFill(start_color='8B6914', end_color='8B6914', fill_type='solid'),
                   2: PatternFill(start_color='6B6B6B', end_color='6B6B6B', fill_type='solid'),
                   3: PatternFill(start_color='7B4B1A', end_color='7B4B1A', fill_type='solid')}

    headers = ['순위', '이름', '학번', '최종 자산 (원)', '수익률 (%)', '수익금액 (원)']
    ws.append(headers)
    for i, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    widths = [8, 12, 14, 18, 13, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    for rank, e in enumerate(board, 1):
        row = [rank, e['name'], e['sid'], e['total_value'], e['gain_pct'], e['gain_amount']]
        ws.append(row)
        r = ws.max_row
        fill = medal_fills.get(rank, up_fill if e['gain_pct'] >= 0 else dn_fill)
        for col, cell in enumerate(ws[r], 1):
            cell.fill = fill
            cell.border = border
            cell.font = Font(color='FFFFFF', bold=(rank <= 3))
            cell.alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')
        ws.cell(r, 4).number_format = '#,##0'
        ws.cell(r, 5).number_format = '0.00'
        ws.cell(r, 6).number_format = '#,##0'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{room.name}_결과.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(
        debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true',
        host='0.0.0.0', port=5000,
    )
